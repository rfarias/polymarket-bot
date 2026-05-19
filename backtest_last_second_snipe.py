"""
backtest_last_second_snipe.py

Backtesting for the Last-Second Snipe strategy.
Uses real Binance historical candles to simulate the 7-indicator signal
and realistic token pricing (no fixed $0.50).

Usage:
  python backtest_last_second_snipe.py --hours 72 --output results.xlsx
  python backtest_last_second_snipe.py --hours 24 --mode safe
  python backtest_last_second_snipe.py --hours 48 --no-excel

Output: Excel with 3 sheets:
  - Summary: 27 configs (9 confidence thresholds × 3 modes)
  - Best Config Trades: individual trade log for best config
  - Bankroll Curves: bankroll evolution for top 5 configs
"""
from __future__ import annotations

import argparse
import json
import math
import sys
import time
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

sys.stdout.reconfigure(encoding="utf-8")

BINANCE_KLINES_URL = "https://api.binance.com/api/v3/klines"


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

@dataclass
class Candle:
    open_time: int   # Unix seconds
    open: float
    high: float
    low: float
    close: float
    volume: float


def fetch_historical_klines(
    symbol: str = "BTCUSDT",
    interval: str = "1m",
    hours: int = 72,
) -> List[Candle]:
    """Fetch historical 1m candles from Binance."""
    limit = min(hours * 60, 1000)
    chunks: List[Candle] = []
    end_ms = int(time.time() * 1000)
    total_needed = hours * 60

    print(f"[Backtest] Fetching {total_needed} × 1m candles from Binance...")
    fetched = 0
    while fetched < total_needed:
        batch = min(1000, total_needed - fetched)
        url = (
            f"{BINANCE_KLINES_URL}?symbol={symbol}&interval={interval}"
            f"&endTime={end_ms}&limit={batch}"
        )
        req = urllib.request.Request(url, headers={"User-Agent": "polymarket-bot/1.0"})
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                rows = json.loads(resp.read())
        except Exception as exc:
            print(f"[Backtest] Fetch error: {exc}")
            break

        if not rows:
            break

        batch_candles = [
            Candle(
                open_time=int(r[0]) // 1000,
                open=float(r[1]),
                high=float(r[2]),
                low=float(r[3]),
                close=float(r[4]),
                volume=float(r[5]),
            )
            for r in rows[:-1]  # drop open candle
        ]
        batch_candles.sort(key=lambda c: c.open_time)
        chunks = batch_candles + chunks
        fetched += len(batch_candles)
        end_ms = rows[0][0] - 1  # move window back
        print(f"  fetched {fetched}/{total_needed}")
        time.sleep(0.1)

    chunks.sort(key=lambda c: c.open_time)
    return chunks


# ---------------------------------------------------------------------------
# Signal computation (no WebSocket — uses only historical candles)
# ---------------------------------------------------------------------------

def _ema(values: List[float], period: int) -> List[float]:
    if not values:
        return []
    k = 2.0 / (period + 1)
    result = [values[0]]
    for v in values[1:]:
        result.append(v * k + result[-1] * (1 - k))
    return result


def _rsi(closes: List[float], period: int = 14) -> float:
    if len(closes) < period + 1:
        return 50.0
    gains, losses = [], []
    for i in range(1, len(closes)):
        d = closes[i] - closes[i - 1]
        gains.append(max(d, 0))
        losses.append(max(-d, 0))
    avg_gain = sum(gains[-period:]) / period
    avg_loss = sum(losses[-period:]) / period
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100.0 - 100.0 / (1 + rs)


def _window_delta_weight(delta_pct: float) -> float:
    d = abs(delta_pct)
    if d >= 0.10:
        return 7.0
    elif d >= 0.02:
        return 5.0
    elif d >= 0.005:
        return 3.0
    elif d >= 0.001:
        return 1.0
    return 0.0


def estimate_token_price(delta_pct: float) -> float:
    d = abs(delta_pct)
    if d < 0.005:
        return 0.50
    elif d < 0.02:
        return 0.50 + (d - 0.005) / (0.02 - 0.005) * 0.05
    elif d < 0.05:
        return 0.55 + (d - 0.02) / (0.05 - 0.02) * 0.10
    elif d < 0.10:
        return 0.65 + (d - 0.05) / (0.10 - 0.05) * 0.15
    elif d < 0.15:
        return 0.80 + (d - 0.10) / (0.15 - 0.10) * 0.12
    else:
        return min(0.97, 0.92 + (d - 0.15) * 0.5)


def compute_signal_historical(
    candles_up_to_t: List[Candle],
    window_open_price: float,
    current_price: float,
) -> Tuple[str, float, float, float]:
    """
    Returns (side, score, confidence, delta_pct) using only candle data.
    No tick feed — uses last candle close as proxy for tick trend.
    """
    closes = [c.close for c in candles_up_to_t]
    volumes = [c.volume for c in candles_up_to_t]
    score = 0.0

    # Ind 1: Window delta
    delta_pct = (current_price - window_open_price) / window_open_price * 100 if window_open_price else 0.0
    w = _window_delta_weight(delta_pct)
    sign = 1.0 if delta_pct >= 0 else -1.0
    score += sign * w

    # Ind 2: Micro momentum (last 2 candles)
    micro = 0.0
    if len(closes) >= 2:
        micro = 1.0 if closes[-1] > closes[-2] else (-1.0 if closes[-1] < closes[-2] else 0.0)
        if len(closes) >= 3:
            micro += 0.5 if closes[-2] > closes[-3] else (-0.5 if closes[-2] < closes[-3] else 0.0)
    score += micro * 2.0

    # Ind 3: Acceleration
    accel = 0.0
    if len(closes) >= 4:
        m1 = closes[-1] - closes[-2]
        m2 = closes[-2] - closes[-3]
        if m1 * m2 > 0:
            accel = (1.0 if abs(m1) > abs(m2) else 0.5) * (1.0 if m1 > 0 else -1.0)
        elif m1 * m2 < 0:
            accel = -0.5 * (1.0 if m1 > 0 else -1.0)
    score += accel * 1.5

    # Ind 4: EMA crossover
    ema_cross = 0.0
    if len(closes) >= 21:
        ema9 = _ema(closes, 9)
        ema21 = _ema(closes, 21)
        ema_cross = 1.0 if ema9[-1] > ema21[-1] else -1.0
    score += ema_cross

    # Ind 5: RSI
    rsi_val = _rsi(closes, 14)
    rsi_contrib = 0.0
    if rsi_val > 75:
        rsi_contrib = 2.0
    elif rsi_val > 55:
        rsi_contrib = 1.0
    elif rsi_val < 25:
        rsi_contrib = -2.0
    elif rsi_val < 45:
        rsi_contrib = -1.0
    score += rsi_contrib

    # Ind 6: Volume surge
    vol_contrib = 0.0
    if len(volumes) >= 6:
        r_avg = sum(volumes[-3:]) / 3
        p_avg = sum(volumes[-6:-3]) / 3
        if p_avg > 0 and r_avg >= p_avg * 1.5:
            vol_contrib = sign
    score += vol_contrib

    # Ind 7: Tick trend proxy (last 3 candle direction)
    tick_proxy = 0.0
    if len(closes) >= 4:
        up_count = sum(1 for i in range(-3, 0) if closes[i] > closes[i - 1])
        dn_count = 3 - up_count
        tick_proxy = (up_count - dn_count) / 3.0
    score += tick_proxy * 2.0

    side = "UP" if score >= 0 else "DOWN"
    confidence = min(abs(score) / 7.0, 1.0)
    return side, score, confidence, delta_pct


# ---------------------------------------------------------------------------
# Backtest engine
# ---------------------------------------------------------------------------

@dataclass
class BacktestTrade:
    window_ts: int
    side: str
    score: float
    confidence: float
    delta_pct: float
    token_price: float
    result: str      # "win" | "loss"
    pnl_usd: float
    bet_size: float
    bankroll_after: float


@dataclass
class BacktestResult:
    config_name: str
    mode: str
    min_confidence: float
    trades: int
    wins: int
    losses: int
    skipped: int
    win_rate: float
    total_pnl: float
    final_bankroll: float
    max_drawdown: float
    trade_log: List[BacktestTrade] = field(default_factory=list)
    bankroll_curve: List[float] = field(default_factory=list)


def _compute_bet(mode: str, bankroll: float, starting: float, min_bet: float) -> float:
    if mode == "degen":
        return max(bankroll, min_bet)
    if mode == "aggressive":
        profit = bankroll - starting
        return max(profit, min_bet) if profit > min_bet else min_bet
    # safe: 25%
    return max(bankroll * 0.25, min_bet)


def run_backtest_config(
    candles: List[Candle],
    starting_bankroll: float,
    min_bet: float,
    mode: str,
    min_confidence: float,
    windows_per_5min: int = 5,  # use candles at T-2min (3 min before close)
) -> BacktestResult:
    """
    Simulate one config across all 5-min windows in the candle history.
    """
    if not candles:
        return BacktestResult(
            config_name=f"{mode}@{min_confidence:.0%}",
            mode=mode, min_confidence=min_confidence,
            trades=0, wins=0, losses=0, skipped=0,
            win_rate=0, total_pnl=0, final_bankroll=starting_bankroll,
            max_drawdown=0,
        )

    bankroll = starting_bankroll
    peak_bankroll = starting_bankroll
    max_drawdown = 0.0
    trades, wins, losses, skipped = 0, 0, 0, 0
    trade_log: List[BacktestTrade] = []
    bankroll_curve = [starting_bankroll]

    # Group candles into 5-min windows
    start_ts = candles[0].open_time
    end_ts = candles[-1].open_time
    window_ts = start_ts - (start_ts % 300)

    # Build index for fast candle lookup
    candle_by_ts: Dict[int, int] = {c.open_time: i for i, c in enumerate(candles)}

    while window_ts + 300 <= end_ts:
        close_ts = window_ts + 300
        # Window open price = price at window_ts (closest 1m candle open)
        wts_aligned = window_ts - (window_ts % 60)
        idx = candle_by_ts.get(wts_aligned)
        if idx is None:
            window_ts += 300
            continue
        window_open_price = candles[idx].open

        # Signal computed at T-2min (index at window_ts + 180 ≈ 3 min in)
        signal_ts = window_ts + 180
        s_aligned = signal_ts - (signal_ts % 60)
        sidx = candle_by_ts.get(s_aligned, idx)
        # Use candles up to signal point
        hist = candles[max(0, sidx - 29):sidx + 1]
        if not hist:
            window_ts += 300
            continue

        current_price = hist[-1].close
        side, score, confidence, delta_pct = compute_signal_historical(
            hist, window_open_price, current_price
        )

        # Skip if confidence below threshold
        if confidence < min_confidence:
            skipped += 1
            window_ts += 300
            continue

        # Token price at this delta
        token_price = estimate_token_price(delta_pct)

        # Determine outcome: was BTC above open at close?
        close_idx = candle_by_ts.get(close_ts - 60)
        if close_idx is None:
            window_ts += 300
            continue
        close_price = candles[close_idx].close
        actual = "UP" if close_price >= window_open_price else "DOWN"
        result = "win" if actual == side else "loss"

        # Compute PnL
        bet = _compute_bet(mode, bankroll, starting_bankroll, min_bet)
        bet = min(bet, bankroll)  # can't bet more than we have
        shares = bet / token_price

        if result == "win":
            pnl = shares * (1.0 - token_price)
            bankroll += pnl
            wins += 1
        else:
            pnl = -bet
            bankroll = max(bankroll + pnl, 0.0)
            losses += 1

        trades += 1
        peak_bankroll = max(peak_bankroll, bankroll)
        drawdown = (peak_bankroll - bankroll) / peak_bankroll if peak_bankroll > 0 else 0
        max_drawdown = max(max_drawdown, drawdown)
        bankroll_curve.append(round(bankroll, 4))

        trade_log.append(BacktestTrade(
            window_ts=window_ts,
            side=side,
            score=round(score, 3),
            confidence=round(confidence, 3),
            delta_pct=round(delta_pct, 6),
            token_price=round(token_price, 4),
            result=result,
            pnl_usd=round(pnl, 4),
            bet_size=round(bet, 4),
            bankroll_after=round(bankroll, 4),
        ))

        if bankroll < min_bet:
            bankroll = starting_bankroll  # reset in backtest

        window_ts += 300

    win_rate = wins / trades if trades > 0 else 0.0
    return BacktestResult(
        config_name=f"{mode}@{min_confidence:.0%}",
        mode=mode,
        min_confidence=min_confidence,
        trades=trades,
        wins=wins,
        losses=losses,
        skipped=skipped,
        win_rate=win_rate,
        total_pnl=round(bankroll - starting_bankroll, 4),
        final_bankroll=round(bankroll, 4),
        max_drawdown=round(max_drawdown, 4),
        trade_log=trade_log,
        bankroll_curve=bankroll_curve,
    )


# ---------------------------------------------------------------------------
# Excel export (Requisito 12)
# ---------------------------------------------------------------------------

def export_excel(results: List[BacktestResult], output_path: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        print("[Backtest] openpyxl not installed — skipping Excel export")
        return

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ----
    ws1 = wb.active
    ws1.title = "Summary"
    headers = [
        "Config", "Mode", "Min Confidence", "Trades", "Wins", "Losses",
        "Skipped", "Win Rate", "Total PnL ($)", "Final Bankroll ($)", "Max Drawdown"
    ]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")

    results_sorted = sorted(results, key=lambda r: r.total_pnl, reverse=True)
    for r in results_sorted:
        ws1.append([
            r.config_name, r.mode, f"{r.min_confidence:.0%}",
            r.trades, r.wins, r.losses, r.skipped,
            f"{r.win_rate:.1%}", round(r.total_pnl, 2),
            round(r.final_bankroll, 2), f"{r.max_drawdown:.1%}",
        ])

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 16

    # ---- Sheet 2: Best Config Trades ----
    ws2 = wb.create_sheet("Best Config Trades")
    best = results_sorted[0] if results_sorted else None
    if best and best.trade_log:
        trade_headers = [
            "Window TS", "Side", "Score", "Confidence", "Delta%",
            "Token Price", "Bet ($)", "Shares", "Result", "PnL ($)", "Bankroll After ($)"
        ]
        ws2.append(trade_headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="2E7D32")
            cell.font = Font(bold=True, color="FFFFFF")
        for t in best.trade_log:
            shares = round(t.bet_size / max(t.token_price, 0.01), 2)
            ws2.append([
                t.window_ts, t.side, t.score, f"{t.confidence:.0%}",
                f"{t.delta_pct:+.4f}%", t.token_price, t.bet_size,
                shares, t.result, t.pnl_usd, t.bankroll_after,
            ])
        for row in ws2.iter_rows(min_row=2):
            if row[8].value == "loss":
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
            elif row[8].value == "win":
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="CCFFCC")

    # ---- Sheet 3: Bankroll Curves (top 5 configs) ----
    ws3 = wb.create_sheet("Bankroll Curves")
    top5 = results_sorted[:5]
    col = 1
    for r in top5:
        ws3.cell(row=1, column=col, value=r.config_name).font = Font(bold=True)
        for i, val in enumerate(r.bankroll_curve):
            ws3.cell(row=i + 2, column=col, value=val)
        col += 1

    # Add line chart
    if top5:
        chart = LineChart()
        chart.title = "Bankroll Curves — Top 5 Configs"
        chart.y_axis.title = "Bankroll ($)"
        chart.x_axis.title = "Trade #"
        max_rows = max(len(r.bankroll_curve) for r in top5)
        for i, r in enumerate(top5):
            data_ref = Reference(ws3, min_col=i + 1, min_row=1, max_row=max_rows + 1)
            chart.add_data(data_ref, titles_from_data=True)
        chart.width = 25
        chart.height = 14
        ws3.add_chart(chart, "G2")

    wb.save(output_path)
    print(f"[Backtest] Excel saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

MODES = ["safe", "aggressive", "degen"]
CONFIDENCE_THRESHOLDS = [0.0, 0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50]


def main() -> None:
    parser = argparse.ArgumentParser(description="Backtest Last-Second Snipe")
    parser.add_argument("--hours", type=int, default=72, help="Hours of history to backtest")
    parser.add_argument("--output", default="results.xlsx", help="Output Excel file")
    parser.add_argument("--mode", choices=MODES + ["all"], default="all",
                        help="Mode to test (default: all 3)")
    parser.add_argument("--bankroll", type=float, default=10.0, help="Starting bankroll")
    parser.add_argument("--min-bet", type=float, default=1.0, help="Min bet size")
    parser.add_argument("--no-excel", action="store_true", help="Skip Excel export")
    args = parser.parse_args()

    # Fetch historical candles
    candles = fetch_historical_klines("BTCUSDT", "1m", args.hours)
    if not candles:
        print("[Backtest] No candles fetched — aborting")
        sys.exit(1)
    print(f"[Backtest] {len(candles)} candles loaded  ({args.hours}h history)")

    modes_to_test = MODES if args.mode == "all" else [args.mode]
    configs = [(m, c) for m in modes_to_test for c in CONFIDENCE_THRESHOLDS]
    print(f"[Backtest] Testing {len(configs)} configs...\n")

    results: List[BacktestResult] = []
    for i, (mode, conf) in enumerate(configs):
        r = run_backtest_config(
            candles=candles,
            starting_bankroll=args.bankroll,
            min_bet=args.min_bet,
            mode=mode,
            min_confidence=conf,
        )
        results.append(r)
        print(
            f"  [{i+1:2d}/{len(configs)}] {r.config_name:22s}  "
            f"trades={r.trades:3d}  WR={r.win_rate:.0%}  "
            f"PnL={r.total_pnl:+.2f}  drawdown={r.max_drawdown:.0%}"
        )

    # Print top 5
    top5 = sorted(results, key=lambda r: r.total_pnl, reverse=True)[:5]
    print(f"\n{'='*60}")
    print("TOP 5 CONFIGS BY PnL:")
    print(f"{'='*60}")
    for r in top5:
        print(
            f"  {r.config_name:22s}  trades={r.trades:3d}  WR={r.win_rate:.0%}"
            f"  PnL={r.total_pnl:+.2f}  bankroll=${r.final_bankroll:.2f}"
            f"  drawdown={r.max_drawdown:.0%}"
        )

    if not args.no_excel:
        export_excel(results, args.output)


if __name__ == "__main__":
    main()
